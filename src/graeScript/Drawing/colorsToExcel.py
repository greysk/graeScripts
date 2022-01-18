import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

from graeScript import outfile_path
from graeScript.Drawing import HtmlColors


def colors2excel(font_name: str = "Cascadia Code", head_bold: bool = True):
    """
    Create a new excel workbook from existing one.
    """
    new_wb_name = outfile_path() / "HTMLcolors.xlsx"
    # Get html color values from database.
    htmlcolors = HtmlColors().all
    # Get column headings for each column in the row.
    headings_by_col = {i: h for i, h in enumerate(htmlcolors[0].keys(), 1)}
    # Add a 'Sample' column which will be colored to match each row's color.
    headings_by_col.setdefault(len(htmlcolors[0]) + 1, 'Sample', )

    # Open new, blank Workbook and obtain the first Worksheet.
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Write rows from database table into the blank Worksheet.
    for i, row in enumerate(htmlcolors, 1):
        # For current row, create new cell object for each column.
        new_cells_row = {heading:
                         sheet.cell(row=i, column=c)
                         for c, heading in headings_by_col.items()
                         }
        # Write Worksheet heading row.
        if i == 1:
            for heading, cell in new_cells_row.items():
                cell.font = Font(name=font_name, bold=head_bold)
                cell.value = heading
            continue
        # Write values into Worksheet from database for non-header rows.
        for heading, cell in new_cells_row.items():
            if heading != 'sample':
                cell.font = Font(name=font_name)
                cell.value = row[heading]
                continue
            # Set fill color of 'Sample' row to the row's color.
            cell.fill = PatternFill(fill_type='solid',
                                    start_color=row['hex'].lstrip('#'),
                                    end_color=row['hex'].lstrip('#'))
    # Save workbook to a file.
    wb.save(new_wb_name)
